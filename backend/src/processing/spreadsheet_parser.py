from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from uuid import NAMESPACE_URL, uuid5

from src.processing.types import BlockType, DependencyUnavailableError, ParsedBlock, ParsedDocument, ParsedPage


SUPPORTED_SPREADSHEET_EXTENSIONS = {"csv", "xlsx", "xls"}

# Header detection: a header row needs ≥ min(3, max_width) cells filled
# and fill_ratio ≥ _HEADER_MIN_FILL.  This skips merged title rows (1 cell).
_HEADER_MIN_FILL = 0.40
# Data rows in a tabular sheet must have ≥ 50% of columns filled.
_TABLE_DATA_FILL = 0.50
_TABLE_MIN_DATA = 2       # minimum qualifying data rows to call a sheet "tabular"
_NUMERIC_RATIO = 0.70     # fraction of non-empty values required to label a column numeric
_MAX_HEADER_SCAN = 10     # scan at most this many rows to locate the header
_MAX_PREAMBLE_ROWS = 5    # keep leading title/description rows as sheet context


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", "").replace("%", "").replace(" ", ""))
        return True
    except ValueError:
        return False


def _format_numeric(value: str) -> str:
    """Strip trailing '.0' from integer-valued floats ('1.0' → '1')."""
    try:
        f = float(value.replace(",", "").replace("%", "").replace(" ", ""))
        if f == int(f) and "%" not in value:
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return value


class SpreadsheetParser:
    def __init__(
        self,
        *,
        max_rows_per_block: int = 40,
        max_columns: int = 60,
        max_cell_chars: int = 500,
        max_verbalized_rows: int = 500,
    ) -> None:
        self.max_rows_per_block = max_rows_per_block
        self.max_columns = max_columns
        self.max_cell_chars = max_cell_chars
        self.max_verbalized_rows = max_verbalized_rows

    def parse(self, file_path: Path, *, language: str = "unknown", display_name: str | None = None) -> ParsedDocument:
        extension = file_path.suffix.lower().lstrip(".")
        if extension not in SUPPORTED_SPREADSHEET_EXTENSIONS:
            raise ValueError(f"SpreadsheetParser does not support .{extension}")

        sheets = (
            self._read_csv(file_path, display_name=display_name)
            if extension == "csv"
            else self._read_workbook(file_path, extension)
        )
        pages: list[ParsedPage] = []
        warnings: list[str] = []

        for page_number, (sheet_name, rows) in enumerate(sheets, start=1):
            clean_rows, row_numbers = self._trim_table(rows)
            if not clean_rows:
                warnings.append(f"Sheet {sheet_name!r} is empty")
                pages.append(ParsedPage(page_number=page_number, blocks=[]))
                continue

            max_width = max(len(r) for r in clean_rows)
            if self._is_tabular(clean_rows, max_width):
                blocks = self._blocks_from_rows(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    rows=clean_rows,
                    row_numbers=row_numbers,
                    page_number=page_number,
                    language=language,
                )
            else:
                # Pass raw (un-trimmed) rows so blank lines work as section separators.
                blocks = self._blocks_from_document(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    raw_rows=rows,
                    page_number=page_number,
                    language=language,
                )
            pages.append(ParsedPage(page_number=page_number, blocks=blocks))

        return ParsedDocument(
            source_path=str(file_path),
            file_type=extension,
            language=language,
            pages=pages or [ParsedPage(page_number=1, blocks=[])],
            warnings=warnings,
            extra={"parser": "spreadsheet", "sheet_count": len(sheets)},
        )

    # ── Sheet reading ────────────────────────────────────────────────────────────

    def _read_csv(self, file_path: Path, *, display_name: str | None = None) -> list[tuple[str, list[list[str]]]]:
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel
            rows = [[self._cell_to_text(cell) for cell in row] for row in csv.reader(handle, dialect)]
        sheet_label = Path(display_name).stem if display_name else file_path.stem
        return [(sheet_label, rows)]

    def _read_workbook(self, file_path: Path, extension: str) -> list[tuple[str, list[list[str]]]]:
        if extension == "xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise DependencyUnavailableError("openpyxl is required for XLSX parsing") from exc
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheets: list[tuple[str, list[list[str]]]] = []
            for worksheet in workbook.worksheets:
                rows = [
                    [self._cell_to_text(cell) for cell in row[: self.max_columns]]
                    for row in worksheet.iter_rows(values_only=True)
                ]
                sheets.append((worksheet.title, rows))
            workbook.close()
            return sheets

        try:
            import xlrd
        except ImportError as exc:
            raise DependencyUnavailableError("xlrd is required for XLS parsing") from exc
        workbook = xlrd.open_workbook(str(file_path), on_demand=True)
        sheets = []
        for sheet in workbook.sheets():
            rows = [
                [
                    self._cell_to_text(sheet.cell_value(row_index, col_index))
                    for col_index in range(min(sheet.ncols, self.max_columns))
                ]
                for row_index in range(sheet.nrows)
            ]
            sheets.append((sheet.name, rows))
        workbook.release_resources()
        return sheets

    # ── Sheet classification ─────────────────────────────────────────────────────

    def _find_header_row_index(self, rows: list[list[str]], max_width: int) -> int:
        """Return the index of the first row that looks like a column header row.

        A title row has only 1 merged cell; a real header has ≥ min(3, max_width) cells
        filled and a fill ratio ≥ _HEADER_MIN_FILL.  Returns -1 if none found.
        """
        if not rows:
            return -1
        min_cells = min(3, max_width)
        for i, row in enumerate(rows[:_MAX_HEADER_SCAN]):
            filled = sum(1 for c in row if c.strip())
            if filled <= 1:
                # Skip merged title rows and short sheet preambles before the true header.
                continue
            if filled >= min_cells and filled / max(1, max_width) >= _HEADER_MIN_FILL:
                return i
        return -1

    def _is_tabular(self, rows: list[list[str]], max_width: int) -> bool:
        """Return True if this sheet has structured column data (not a document-style sheet)."""
        if max_width < 2:
            return False
        header_idx = self._find_header_row_index(rows, max_width)
        if header_idx < 0:
            return False
        data_rows = rows[header_idx + 1 :]
        if len(data_rows) < _TABLE_MIN_DATA:
            return False
        min_cells = max(2, int(max_width * _TABLE_DATA_FILL))
        qualifying = sum(1 for r in data_rows if sum(1 for c in r if c.strip()) >= min_cells)
        return qualifying >= _TABLE_MIN_DATA

    # ── Column helpers ───────────────────────────────────────────────────────────

    def _infer_column_types(self, data_rows: list[list[str]], header: list[str]) -> list[str]:
        """Classify each column as 'numeric' or 'text' based on sampled data values."""
        types: list[str] = []
        for col_idx in range(len(header)):
            vals = [
                row[col_idx].strip()
                for row in data_rows
                if col_idx < len(row) and row[col_idx].strip()
            ]
            if not vals:
                types.append("text")
                continue
            numeric_count = sum(1 for v in vals if _looks_numeric(v))
            types.append("numeric" if numeric_count / len(vals) >= _NUMERIC_RATIO else "text")
        return types

    # ── Tabular sheet → blocks ───────────────────────────────────────────────────

    def _blocks_from_rows(
        self,
        *,
        file_path: Path,
        sheet_name: str,
        rows: list[list[str]],
        row_numbers: list[int],
        page_number: int,
        language: str,
    ) -> list[ParsedBlock]:
        max_width = max(len(r) for r in rows)
        header_idx = self._find_header_row_index(rows, max_width)
        if header_idx < 0:
            header_idx = 0
        header = self._header_for(rows[header_idx])
        data_rows = rows[header_idx + 1 :]
        col_types = self._infer_column_types(data_rows, header)
        preamble = self._extract_preamble(rows[:header_idx])
        blocks: list[ParsedBlock] = []

        summary_block = self._build_summary_block(
            file_path=file_path,
            sheet_name=sheet_name,
            header=header,
            col_types=col_types,
            row_count=len(data_rows),
            page_number=page_number,
            language=language,
            block_index=0,
            preamble=preamble,
        )
        if summary_block is not None:
            blocks.append(summary_block)

        # Markdown table blocks (batched to keep chunks manageable)
        for start in range(0, len(data_rows), self.max_rows_per_block):
            row_batch = data_rows[start : start + self.max_rows_per_block]
            markdown = self._to_markdown(header, row_batch)
            if not markdown.strip():
                continue
            row_start = row_numbers[header_idx + 1 + start]
            row_end = row_numbers[header_idx + 1 + start + len(row_batch) - 1]
            blocks.append(
                ParsedBlock(
                    block_id=self._block_id(file_path, sheet_name, row_start, row_end, markdown),
                    block_index=len(blocks),
                    block_type=BlockType.TABLE.value,
                    content=markdown,
                    page_number=page_number,
                    language=language,
                    reading_order=len(blocks),
                    source="spreadsheet",
                    extra={
                        "block_kind": "table_block",
                        "sheet_name": sheet_name,
                        "row_start": row_start,
                        "row_end": row_end,
                        "columns": header,
                    },
                )
            )

        # One verbalized sentence per row — enables keyword / QA retrieval per record
        if data_rows and len(data_rows) <= self.max_verbalized_rows:
            for index, row in enumerate(data_rows):
                actual_row_index = row_numbers[header_idx + 1 + index]
                verbalized = self._verbalize_row(
                    header=header,
                    row=row,
                    col_types=col_types,
                    sheet_name=sheet_name,
                    row_index=actual_row_index,
                    language=language,
                )
                if not verbalized:
                    continue
                blocks.append(
                    ParsedBlock(
                        block_id=self._block_id(
                            file_path,
                            sheet_name,
                            actual_row_index,
                            actual_row_index,
                            f"row::{verbalized[:80]}",
                        ),
                        block_index=len(blocks),
                        block_type=BlockType.PARAGRAPH.value,
                        content=verbalized,
                        page_number=page_number,
                        language=language,
                        reading_order=len(blocks),
                        source="spreadsheet",
                        extra={
                            "block_kind": "table_row",
                            "sheet_name": sheet_name,
                            "row_index": actual_row_index,
                            "columns": header,
                            "col_types": col_types,
                        },
                    )
                )

        return blocks

    # ── Document-style sheet → blocks ────────────────────────────────────────────

    def _blocks_from_document(
        self,
        *,
        file_path: Path,
        sheet_name: str,
        raw_rows: list[list[str]],
        page_number: int,
        language: str,
    ) -> list[ParsedBlock]:
        """Extract content from a sparse / document-style sheet (no structured columns).

        Blank rows act as section separators; each section becomes one PARAGRAPH block.
        The first single-cell section (sheet title) is skipped.
        """
        sections = self._split_into_sections(raw_rows)
        blocks: list[ParsedBlock] = []
        first_content = True

        for section_rows in sections:
            non_empty = [r for r in section_rows if any(c.strip() for c in r)]
            if not non_empty:
                continue

            # Skip the sheet-level title: first section that is a single short cell.
            if first_content:
                first_content = False
                all_cells = [c.strip() for r in non_empty for c in r if c.strip()]
                if len(all_cells) == 1 and len(non_empty) == 1:
                    continue

            content = self._section_to_text(non_empty)
            if not content.strip():
                continue

            blocks.append(
                ParsedBlock(
                    block_id=self._block_id(file_path, sheet_name, len(blocks), len(blocks), content),
                    block_index=len(blocks),
                    block_type=BlockType.PARAGRAPH.value,
                    content=content,
                    page_number=page_number,
                    language=language,
                    reading_order=len(blocks),
                    source="spreadsheet",
                    extra={
                        "block_kind": "document_section",
                        "sheet_name": sheet_name,
                    },
                )
            )

        return blocks

    @staticmethod
    def _split_into_sections(rows: list[list[str]]) -> list[list[list[str]]]:
        """Group rows into sections using blank rows as separators."""
        sections: list[list[list[str]]] = []
        current: list[list[str]] = []
        for row in rows:
            if not any(c.strip() for c in row):
                if current:
                    sections.append(current)
                    current = []
            else:
                current.append(row)
        if current:
            sections.append(current)
        return sections

    @staticmethod
    def _section_to_text(rows: list[list[str]]) -> str:
        """Render a section's rows as human-readable text suitable for RAG indexing."""
        lines: list[str] = []
        for row in rows:
            cells = [c.strip() for c in row if c.strip()]
            if not cells:
                continue
            if len(cells) == 1:
                lines.append(cells[0])
            elif len(cells) == 2 and re.match(r"^\d+$", cells[0]):
                # Numbered list item: "1 | Question text" → "1. Question text"
                lines.append(f"{cells[0]}. {cells[1]}")
            else:
                lines.append(": ".join(cells))
        return "\n".join(lines)

    # ── Summary block ────────────────────────────────────────────────────────────

    def _build_summary_block(
        self,
        *,
        file_path: Path,
        sheet_name: str,
        header: list[str],
        col_types: list[str],
        row_count: int,
        page_number: int,
        language: str,
        block_index: int,
        preamble: list[str] | None = None,
    ) -> ParsedBlock | None:
        if not header:
            return None
        col_parts: list[str] = []
        for name, ctype in zip(header, col_types):
            label = name.strip() if name.strip() else "—"
            col_parts.append(f"{label} (số)" if ctype == "numeric" else label)
        column_list = ", ".join(col_parts)
        preamble_text = " | ".join(preamble or [])
        if language == "en":
            prefix = f"Preamble: {preamble_text}. " if preamble_text else ""
            content = (
                f"{prefix}Table summary — sheet '{sheet_name}' has {row_count} rows × {len(header)} columns. "
                f"Columns: {column_list}."
            )
        else:
            prefix = f"Tiền đề: {preamble_text}. " if preamble_text else ""
            content = (
                f"{prefix}Tóm tắt bảng — sheet '{sheet_name}' có {row_count} hàng × {len(header)} cột. "
                f"Các cột: {column_list}."
            )
        return ParsedBlock(
            block_id=self._block_id(file_path, sheet_name, 0, 0, f"summary::{content[:80]}"),
            block_index=block_index,
            block_type=BlockType.PARAGRAPH.value,
            content=content,
            page_number=page_number,
            language=language,
            reading_order=block_index,
            source="spreadsheet",
            extra={
                "block_kind": "table_summary",
                "sheet_name": sheet_name,
                "columns": header,
                "col_types": col_types,
                "row_count": row_count,
                "preamble": preamble or [],
            },
        )

    # ── Row verbalization ────────────────────────────────────────────────────────

    def _verbalize_row(
        self,
        *,
        header: list[str],
        row: list[str],
        col_types: list[str] | None,
        sheet_name: str,
        row_index: int,
        language: str,
    ) -> str:
        width = len(header)
        cells = list((row + [""] * width)[:width])
        types = list(((col_types or []) + ["text"] * width)[:width])
        pairs: list[str] = []
        for col_name, cell_value, col_type in zip(header, cells, types):
            value = cell_value.strip()
            if not value:
                continue
            label = col_name.strip() or "—"
            display = _format_numeric(value) if col_type == "numeric" else value
            pairs.append(f"{label}: {display}")
        if not pairs:
            return ""
        body = ". ".join(pairs)
        if language == "en":
            return f"Row {row_index} of table '{sheet_name}'. {body}."
        return f"Hàng {row_index} của bảng '{sheet_name}'. {body}."

    # ── Table helpers ────────────────────────────────────────────────────────────

    def _trim_table(self, rows: list[list[str]]) -> tuple[list[list[str]], list[int]]:
        indexed_rows = [
            (row_index + 1, row[: self.max_columns])
            for row_index, row in enumerate(rows)
            if any(cell.strip() for cell in row)
        ]
        if not indexed_rows:
            return [], []
        row_numbers = [row_number for row_number, _ in indexed_rows]
        non_empty_rows = [row for _, row in indexed_rows]
        max_width = max(len(row) for row in non_empty_rows)
        normalized = [row + [""] * (max_width - len(row)) for row in non_empty_rows]
        non_empty_columns = [
            index
            for index in range(max_width)
            if any(row[index].strip() for row in normalized)
        ][: self.max_columns]
        return [[row[index] for index in non_empty_columns] for row in normalized], row_numbers

    def _header_for(self, row: list[str]) -> list[str]:
        return [cell if cell else f"Column {index + 1}" for index, cell in enumerate(row[: self.max_columns])]

    def _extract_preamble(self, rows: list[list[str]]) -> list[str]:
        """Collect leading title/description rows to keep them as sheet context.

        These rows are preserved in the summary block but never used as column headers.
        """
        preamble: list[str] = []
        for row in rows[:_MAX_PREAMBLE_ROWS]:
            cells = [cell.strip() for cell in row if cell.strip()]
            if not cells:
                continue
            preamble.append(" | ".join(cells))
        return preamble

    def _to_markdown(self, header: list[str], rows: list[list[str]]) -> str:
        if not rows:
            return ""
        width = len(header)
        safe_header = [self._escape_markdown_cell(cell) for cell in header]
        lines = [
            "| " + " | ".join(safe_header) + " |",
            "| " + " | ".join("---" for _ in safe_header) + " |",
        ]
        for row in rows:
            cells = [self._escape_markdown_cell(cell) for cell in (row + [""] * width)[:width]]
            lines.append("| " + " | ".join(cells) + " |")
        return "\n".join(lines)

    def _cell_to_text(self, value) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        text = str(value).replace("\x00", " ").strip()
        return " ".join(text.split())[: self.max_cell_chars]

    @staticmethod
    def _escape_markdown_cell(value: str) -> str:
        return value.replace("|", "\\|").replace("\n", " ").strip()

    @staticmethod
    def _block_id(file_path: Path, sheet_name: str, row_start: int, row_end: int, content: str) -> str:
        return f"blk-{uuid5(NAMESPACE_URL, f'{file_path}:{sheet_name}:{row_start}:{row_end}:{content[:80]}').hex[:12]}"
